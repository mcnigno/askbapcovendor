import os
import openpyxl
from app import db
from app.models import Document
from flask_appbuilder.models.sqla.interface import SQLAInterface


def vendorList():
    file = open('check/lista_cod_vendor.xlsx', mode='rb')
    book = openpyxl.load_workbook(file)
    sheet = book.active
    vendorList = {}
    for row in sheet.iter_rows(min_row=2):
        
        vendorList[row[1].value] = (row[2].value,row[3].value)
    return vendorList

VDL = vendorList()

def update_from_xlsx(file):
    session = db.session
    print('update FUNCTION!')
    book = openpyxl.load_workbook(file)
    sheet = book.active
    a1 = sheet['A6']
    print(a1.value)
    
    reserved_list = []
    updated_list = []
    for row in sheet.iter_rows(min_row=6):
        bapco_code = row[4].value
        oldcode = row[3].value
        print(bapco_code, oldcode)
        doc = db.session.query(Document).filter(Document.code == str(bapco_code)).first()

        if doc and doc.oldcode == 'empty':
            print('this is the ID' ,doc.id)
            datamodel = SQLAInterface(Document, session=session)
            print('BEFORE oldcode', doc.oldcode)
            doc.oldcode = oldcode
            updated_list.append([doc.id, doc.code, doc.oldcode])
            datamodel.edit(doc)

        else:
            reserved_list.append([bapco_code, oldcode])
    return reserved_list, updated_list

found_list = []
notfound_list = []
wrong_code = []
wrong_technip_code = []

def checkMr(file):
    book = openpyxl.load_workbook(file)
    sheet = book.active
    
    for row in sheet.iter_rows(min_row=6):
        
        bapco_code = str(row[4].value or '').strip()
        t_code = str(row[3].value or '').strip()
        mr_code = str(row[2].value or '').strip()

        try:
            if bapco_code.split('-')[4] == '001':
                
                try:
                    #print('VDL:',VDL[bapco_code], 'MR:', bapco_code, t_code, mr_code)
                    found_list.append((bapco_code, VDL[bapco_code], (t_code, mr_code)))
                    #_ = book.active.cell(row=row,column=1,value=5)
                    vdl_code = VDL[bapco_code][0]
                    try:
                        vdl_code = vdl_code.split(' ')[0]
                    except:pass
                    try:
                        t_code = t_code.split(' ')[0]
                    except:pass
                    
                    if vdl_code != t_code:
                        
                        print('Different Technip code:     ',vdl_code,'   ---->   ',t_code )
                        
                        wrong_technip_code.append([bapco_code, mr_code, str(vdl_code), str(t_code)])
                    
                except:
                    print('NOT FOUND IN VDL', bapco_code, t_code, mr_code, file)
                    notfound_list.append((bapco_code, t_code, mr_code))
                
            '''
            else:    
                print('skip sheet:',bapco_code)
            '''
        except:
            #print('   WRONG:   ', bapco_code)
            wrong_code.append([bapco_code,t_code, file])
    
    return found_list, notfound_list

def batch_xlsx():
    for file in os.listdir("app/tmpdir"):
        if file.endswith(".xlsx"):
            print(os.path.join("app/tmpdir", file))
            file = open(os.path.join("app/tmpdir", file), mode='rb')
            print('update from xlsx start')
            #checkMr(file)
            checkMr(os.path.join(file.name))
    print('update from xlsx finish')
    
    print('Wrong Code found List:', wrong_code)
    print('Not found list')
    print(notfound_list)
    print('Wrong Technip Code')
    #print([print(x) for x in wrong_technip_code])
    write_result()

def write_result():
    result = openpyxl.Workbook()
    
    # Contractor Code
    rsheet = result.create_sheet('Contractor_Code')
    label = ['Bapco Code','Material Requisition', 'AskBapco','Vendor List']
    rsheet.append(label)
    for row, x in enumerate(wrong_technip_code):
        rsheet.append(x)

    # Wrong Bapco Code
    rsheet = result.create_sheet('VDL BapcoCode')
    label = ['Wrong Bapco Code','Contractor Code', 'File']
    rsheet.append(label)
    for row, x in enumerate(wrong_code):
        rsheet.append(x)

    # VDL Not Found in ASKBapco
    rsheet = result.create_sheet('VDL NotFound')
    label = ['Wrong Bapco Code', 'File']
    rsheet.append(label)
    for row, x in enumerate(notfound_list):
        rsheet.append(x)
    
    result.save('result.xlsx')




batch_xlsx()

def checkMr(file):
    book = openpyxl.load_workbook(file)
    sheet = book.active
    
    for row in sheet.iter_rows(min_row=6):
        bapco_code = row[4].value
        t_code = row[3].value
        mr_code = row[2].value

        try:
            print('VDL:',VDL[bapco_code], 'MR:', bapco_code, t_code, mr_code)
        except:
            print('NOT FOUND IN VDL', bapco_code, t_code, mr_code)
